from itertools import islice
from pathlib import Path
from tqdm import tqdm, tqdm_pandas
import argparse
import openpyxl
import pandas as pd


tqdm.pandas()


def get_args():
    parser = argparse.ArgumentParser()

    parser.add_argument("--dir", required=False)
    parser.add_argument("--target", required=False)
    parser.add_argument("--spoken_or_written", required=False)
    parser.add_argument("--skip_validation", action="store_true", default=False)
    parser.add_argument("--sim_thresh", type=float, default=0.9, required=False)
    parser.add_argument(
        "--sim_first_time",
        action="store_true",
        required=False,
        help="When it's first time to check similarities",
    )
    parser.add_argument("--runs_mt", action="store_true", required=False)
    parser.add_argument("--mt_similarity", action="store_true", required=False)
    parser.add_argument("--chunk_size", type=int, default=30000, required=False)
    parser.add_argument("--word", required=False)
    parser.add_argument("--category", required=False)

    args = parser.parse_args()
    return args


def convert_ws_to_df(ws) -> pd.DataFrame:
    data = ws.values
    cols = next(data)
    data = list(data)
    data = (islice(r, 0, None) for r in data)
    return pd.DataFrame(data, columns=cols)


def revive_prefix_quote(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    for row in ws.iter_cols(min_col=4, max_col=4):
        for i in range(ws.max_row):
            cell = row[i]
            if cell.quotePrefix:
                cell.value = "'" + str(cell.value)

    df = convert_ws_to_df(ws)

    if None in df.columns.tolist():
        df.drop([None], axis=1, inplace=True)

    if "translation id" in df.columns.tolist():
        df.dropna(subset=["translation id"], inplace=True)
    elif "ko_id" in df.columns.tolist():
        df.dropna(subset=["ko_id"], inplace=True)
    return df


def is_valid_quote_string(sentence):
    n_quote = sentence.count("'")
    n_double_quote = sentence.count('"')

    if n_quote % 2 != 0 or n_double_quote % 2 != 0:
        return False
    else:
        return True


def get_languages_from_path(path):
    path = Path(path)

    stem = path.stem
    if stem.split("_", 1)[0].isdigit():
        _, _, langs, _ = stem.split("_", 3)
    else:
        _, langs, _, _ = stem.split("_", 3)

    if len(langs) == 4:
        langs += "A"

    langs = langs[2:]
    lang1 = "ko" if langs[2:] == "A" else "en"
    lang2 = langs[:2].lower()

    if "_Done" in stem:
        lang1 = "ko"
    return lang1, lang2


def split_into_pass_and_fail(df):
    cols_bool = [
        col
        for col in df.columns.tolist()
        if ("_matches_with_" not in col and "_not_modified" not in col)
    ]
    all_cols_true = df[cols_bool].all(axis=1, bool_only=True)
    df_pass = df[all_cols_true]
    df_fail = df[~all_cols_true]
    return df_pass, df_fail


def drop_columns_with_all_rows_true(df):
    for col in df.columns.tolist():
        if pd.api.types.is_bool_dtype(df[col]) and df[col].all():
            df = df.drop([col], axis=1)
    return df
